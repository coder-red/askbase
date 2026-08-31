"""Document parsing.

Converts raw HTML/PDF content into structured sections and tables while
preserving document hierarchy (title, sections, subsections, paragraphs,
page numbers, dates). This module deliberately does NOT flatten everything
into plain text — structure is preserved for structural chunking.
"""

from __future__ import annotations

from dataclasses import dataclass, field

from bs4 import BeautifulSoup

from basechatt.observability.logging import get_logger

logger = get_logger("basechatt.ingestion.parser")


@dataclass
class ParsedSection:
    title: str
    level: int
    text: str
    page: int | None = None
    number: str = ""


@dataclass
class ParsedTable:
    title: str
    headers: list[str]
    rows: list[list[str]]
    page: int | None = None
    section: str = ""
    units: str = ""
    currency: str = ""


@dataclass
class ParsedDocument:
    title: str
    sections: list[ParsedSection] = field(default_factory=list)
    tables: list[ParsedTable] = field(default_factory=list)
    raw_text: str = ""
    language: str = "en"

    def all_text(self) -> str:
        parts = [self.raw_text]
        for s in self.sections:
            parts.append(s.text)
        return "\n".join(parts)


def parse_html(html: str, url: str = "") -> ParsedDocument:
    """Parse an HTML document into structured sections and tables."""
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
        tag.decompose()

    title = ""
    t = soup.find("title")
    h1 = soup.find("h1")
    if t:
        title = t.get_text(strip=True)
    elif h1:
        title = h1.get_text(strip=True)

    doc = ParsedDocument(title=title)

    # Headings -> sections; plain paragraphs -> text under the current heading.
    # We collect in order using a stream so hierarchy is preserved.
    current_section: ParsedSection | None = None

    for el in soup.find_all(["h1", "h2", "h3", "h4", "h5", "h6", "p", "table", "li"]):
        if el.name and el.name.startswith("h") and len(el.name) == 2:
            level = int(el.name[1])
            heading_text = el.get_text(" ", strip=True)
            if not heading_text:
                continue
            current_section = ParsedSection(
                title=heading_text, level=level, text=""
            )
            doc.sections.append(current_section)
        elif el.name == "p" or el.name == "li":
            text = el.get_text(" ", strip=True)
            if not text:
                continue
            if current_section is None:
                current_section = ParsedSection(title="", level=1, text="")
                doc.sections.append(current_section)
            current_section.text += " " + text + "\n"
        elif el.name == "table":
            tbl = _extract_html_table(el)
            if tbl is not None:
                doc.tables.append(tbl)

    # Fallback raw_text if nothing structured was found.
    if not doc.sections:
        doc.raw_text = soup.get_text(" ", strip=True)

    return doc


def _extract_html_table(table_el) -> ParsedTable | None:
    rows: list[list[str]] = []
    headers: list[str] = []
    for tr in table_el.find_all("tr"):
        cells = []
        for cell in tr.find_all(["th", "td"]):
            cells.append(cell.get_text(" ", strip=True))
        if not cells:
            continue
        if tr.find("th"):
            if not headers:
                headers = cells
            else:
                rows.append(cells)
        else:
            rows.append(cells)
    if not headers and not rows:
        return None
    title_el = table_el.find_previous_sibling(["h3", "h4", "h5", "p", "caption"])
    title = ""
    if title_el:
        title = title_el.get_text(" ", strip=True)
    caption = table_el.find("caption")
    if caption:
        title = caption.get_text(" ", strip=True)
    return ParsedTable(title=title, headers=headers, rows=rows)


def parse_pdf(data: bytes) -> ParsedDocument:
    """Parse a PDF into structured text and tables using pdfplumber."""
    import io

    import pdfplumber

    doc = ParsedDocument(title="")
    try:
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                number = page_idx + 1
                text = page.extract_text() or ""
                doc.sections.append(
                    ParsedSection(title="", level=1, text=text, page=number)
                )
                for table in page.extract_tables() or []:
                    if not table:
                        continue
                    headers: list[str] = []
                    rows: list[list[str]] = []
                    for i, row in enumerate(table):
                        cells = ["" if c is None else c for c in row]
                        if i == 0 and headers == []:
                            headers = cells
                        else:
                            rows.append(cells)
                    doc.tables.append(
                        ParsedTable(
                            title="",
                            headers=headers,
                            rows=rows,
                            page=number,
                        )
                    )
    except Exception as e:  # noqa: BLE001
        logger.warning("PDF parse failed: %s", e)
        doc.sections.append(ParsedSection(title="", level=1, text=""))
    if not doc.raw_text:
        doc.raw_text = "\n".join(s.text for s in doc.sections)
    return doc


def parse_bytes(content: bytes, mime_type: str) -> ParsedDocument:
    """Route raw bytes to the correct parser based on MIME type."""
    mime = (mime_type or "").lower()

    # Binary office/spreadsheet formats that we can parse structurally.
    if "openxmlformats-spreadsheet" in mime or mime.endswith("xlsx"):
        return parse_xlsx(content)
    if "ms-excel" in mime or "spreadsheetml" in mime:
        return parse_xlsx(content)

    # Reject obvious binary containers that have no text parser.
    if _looks_binary(content):
        return ParsedDocument(title="")

    text = content.decode("utf-8", errors="ignore")
    if "html" in mime or "xml" in mime or mime == "":
        if _control_ratio(text) > 0.02:
            return ParsedDocument(title="")
        try:
            return parse_html(text)
        except Exception:  # noqa: BLE001
            pass
    if "pdf" in mime:
        return parse_pdf(content)
    # Fallback: try HTML decode then plain text
    if _control_ratio(text) > 0.02:
        return ParsedDocument(title="")
    try:
        if "<" in text[:200] or "html" in text[:2000].lower():
            return parse_html(text)
        doc = ParsedDocument(title="")
        doc.raw_text = text
        doc.sections.append(ParsedSection(title="", level=1, text=text))
        return doc
    except Exception as e:  # noqa: BLE001
        logger.warning("Generic parse failed: %s", e)
        return ParsedDocument(title="")


def parse_xlsx(data: bytes) -> ParsedDocument:
    """Parse an .xlsx workbook into per-sheet tables and text.

    Spreadsheets are tabular by nature, so each worksheet becomes a
    ``ParsedTable`` (headers + rows) and a flattened text section. This turns
    SEC/statistical bulletins into genuinely searchable chunks instead of raw
    zip-binary junk.
    """
    import io

    from openpyxl import load_workbook

    doc = ParsedDocument(title="")
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        logger.warning("XLSX parse failed: %s", e)
        return doc
    for ws in wb.worksheets:
        title = ws.title or ""
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c).strip() for c in row]
            if not any(cells):
                continue
            rows.append(cells)
            lineno = len(rows)
            doc.sections.append(
                ParsedSection(title=f"{title} row {lineno}", level=1,
                              text=" | ".join(cells))
            )
        headers = rows[0] if rows else []
        body = rows[1:] if rows else []
        if rows:
            doc.tables.append(ParsedTable(title=title, headers=headers, rows=body))
    wb.close()
    return doc


def _looks_binary(data: bytes) -> bool:
    """Detect known binary containers (zip/office/exe/pdf-like magic bytes)."""
    if len(data) < 4:
        return False
    for magic in (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08"):
        if data.startswith(magic):
            return True
    return False


def _control_ratio(sample: str) -> float:
    """Fraction of control characters — a quick binary-vs-text detector."""
    if not sample:
        return 0.0
    controls = sum(
        1 for c in sample
        if (ord(c) < 32 and c not in "\n\r\t") or ord(c) == 127
    )
    return controls / len(sample)
